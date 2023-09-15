# -*- coding: utf-8 -*-
"""
@author: jw
@contact: 752911233@qq.com
@time: 2020/4/29 12:54
@file: handle_excel.py
"""
import time
from collections import namedtuple
from openpyxl import load_workbook
from openpyxl.styles import colors, Font
import shutil
from datetime import datetime
import os

current_dir = os.path.dirname(os.path.abspath(__file__))


class HandleExcel(object):
    """定义处理Excel类"""

    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname
        self.wb = load_workbook(self.filename)
        self.ws = self.wb[self.sheetname] if self.sheetname is not None else self.wb.active

        self.sheet_head_tuple = tuple(self.ws.iter_rows(max_row=1, values_only=True))[0]
        self.Cases = namedtuple("Cases", self.sheet_head_tuple)
        self.cases_list = []

    def get_cases(self):
        """获取所有测试用例"""
        for data in self.ws.iter_rows(min_row=2, values_only=True):
            self.cases_list.append(self.Cases(*data))
        return self.cases_list

    def get_case(self, row):
        """获取某一条测试用例"""
        if isinstance(row, int) and (1 <= row <= self.ws.max_row):
            data = tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
            self.cases_list.append(self.Cases(*data))
            return self.cases_list
        elif isinstance(row, list) or isinstance(row, tuple) and (1 <= row[0] <= row[-1] <= self.ws.max_row):
            for data in self.ws.iter_rows(min_row=row[0], max_row=row[-1], values_only=True):
                self.cases_list.append(self.Cases(*data))
            return self.cases_list
        else:
            print("传入的行号有误，请检查参数类型")

    def write_result(self, row, actual, result):
        """测试执行结果写入表格"""
        other_wb = load_workbook(self.filename)
        other_ws = other_wb[self.sheetname]

        if isinstance(row, int) and (2 <= row <= other_ws.max_row):
            if result == "Fail":
                other_ws.cell(row=row, column=13, value=actual).font = Font(color=colors.RED)
                other_ws.cell(row=row, column=14, value=result).font = Font(color=colors.RED)
                other_wb.save(self.filename)
            else:
                other_ws.cell(row=row, column=13, value=actual)
                other_ws.cell(row=row, column=14, value=result)
                other_wb.save(self.filename)
        else:
            print("请检查传入的行号，行号应为大于1的整数")

    @staticmethod
    def copy_file(file_path):
        new_file_name = "cases_result" + datetime.strftime(datetime.now(), "%Y%m%d%H%M") + ".xlsx"
        new_file_path = os.path.join(current_dir, new_file_name)
        shutil.copyfile(file_path, new_file_path)
        return new_file_path


if __name__ == '__main__':
    start_time = time.time()
    test_case_path = os.path.join(current_dir, "fu_testcase.xlsx")
    one_excel = HandleExcel(filename=test_case_path, sheetname="withcoupon")
    cases = one_excel.get_cases()
    print(f"读取数据耗时：{time.time() - start_time}")
    pass
