
# -*- coding: utf-8 -*-

import os
from collections import namedtuple
import openpyxl


class Case:
    # 这个类用来存储用例
    # 这个方法只是 r_data_obj_new 需要
    def __init__(self, attrs):
        """
        通用，减少代码重复
        :param attrs:[(k,v),(),()]
        """
        for item in attrs:
            setattr(self, item[0], item[1])


class ReadExcel(object):
    Case = namedtuple("Case", ["k", "v"])

    # 两种方式，按行和按最大行读取，还可用对象接收
    def __init__(self, file_name, sheet_name):
        """
        这个用来初始化读取对象的
        :param file_name: 文件名字
        :param sheet_name: 表单名字
        """
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        # 打开工作簿
        self.wb = openpyxl.load_workbook(self.file_name)
        # 选择表单
        self.sheet = self.wb[self.sheet_name]

    def close(self):
        self.wb.close()

    # def __del__(self):
    #     self.wb.close()

    def read_data_line(self):
        """
        按行读取数据
        :return:
        """
        # 按行获取数据转化为列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)
        # 定义一个空列表用来存储所有的用例
        cases = []
        for case in rows_data[1:]:
            data = []
            for cell in case:
                # 判断是否为字符串类型，是则使用eval转换
                if isinstance(cell.value, str):
                    data.append(eval(cell.value))
                else:
                    data.append(cell.value)
            cases.append(dict(list(zip(titles, data))))
        return cases

    def read_data_testcase(self):
        self.open()
        # 按行获取数据转化为列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息 12个标题
        titles = [title.value for title in rows_data[0][0:2]] + ["标识"] + [title.value for title in rows_data[0][3:12]]
        Case = namedtuple("Case", titles)
        # 定义一个空列表用来存储所有的用例
        cases = []
        for index in range(len(rows_data[1:])):
            # 创建一个Case类的对象，用来保存用例数据
            case = rows_data[1:][index]
            # 参加促销活动 and 参与多倍积分测试案例 and 根据列表【2,4,6】编号筛选值
            if case[2].value != None:
                data = [case[i].value for i in range(12)]
                case_obj = Case(*data)
                cases.append(case_obj)
        self.close()
        return cases

    def read_data_obj(self, li):
        self.open()
        # 按行获取数据转化为列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = [title.value for title in rows_data[0] if
                  title.value in ["用例名称", "接口地址", "是否登陆", "请求方式", "请求数据", "期望返回状态", "期望返回数据"]]
        # 定义一个空列表用来存储所有的用例
        cases = []
        flag = ""
        for index in range(len(rows_data[1:])):
            # 创建一个Case类的对象，用来保存用例数据
            case = rows_data[1:][index]
            # 参加促销活动 and 参与多倍积分测试案例 and 根据列表【2,4,6】编号筛选值
            condition = 'case[1].value == "1"' if len(
                li) == 0 else 'case[1].value == "1" and int(case[4].value) in li'  # and case[3].value == "1"
            if eval(condition):
                if case[0].value != None:
                    flag = str(case[0].value)
                data = [flag, case[2].value, case[3].value, case[4].value, case[5].value, case[6].value, case[8].value]
                case = list(zip(titles, data))
                case_obj = Case(case)
                cases.append(case_obj)
                # 数据读完马上关闭表单，避免多个用例打开文件，最后回写数据不成功
        self.close()
        return cases

    def read_data_obj_new(self, li):
        self.open()
        # 按行获取数据转化为列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = [title.value for title in rows_data[0] if
                  title.value in ["用例名称", "接口地址", "是否登陆", "请求方式", "请求数据", "期望返回状态", "期望返回数据"]]
        # 定义一个空列表用来存储所有的用例
        cases = []
        flag = ""
        for index in range(len(rows_data[1:])):
            # 创建一个Case类的对象，用来保存用例数据
            case = rows_data[1:][index]
            # 参加促销活动 and 参与多倍积分测试案例 and 根据列表【2,4,6】编号筛选值
            condition = 'case[1].value == "1"' if len(
                li) == 0 else 'case[1].value == "1" and int(case[4].value) in li'  # and case[3].value == "1"
            if eval(condition) and len(eval(case[8].value)["check_data"].keys()) >= 2 and eval(case[8].value)["info_order_confirm"]["total_discount_amount"] > 0.00:
                if case[0].value != None:
                    flag = str(case[0].value)
                data = [flag, case[2].value, case[3].value, case[4].value, case[5].value, case[6].value, case[8].value]
                case = list(zip(titles, data))
                case_obj = Case(case)
                cases.append(case_obj)
                # 数据读完马上关闭表单，避免多个用例打开文件，最后回写数据不成功
        self.close()
        return cases


    def read_data_obj1(self):
        # 读取数据之前打开表单
        self.open()
        """
        按行读取数据 每个用例存储在一个对象中
        :return:
        """
        # 按行获取数据转化为列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = []
        for title in rows_data[0]:
            if None != title.value:
                titles.append(title.value)
        titles = ["index"] + titles
        # 定义一个空列表用来存储所有的用例
        cases = []
        for index in range(len(rows_data[1:])):
            # 创建一个Case类的对象，用来保存用例数据
            case = rows_data[1:][index]
            # 看是否有最左边那一列
            if None == case[0].value:
                case = case[1:]
            # 筛选标识为1的可用用例
            if case[1].value != "1":
                continue
            data = []
            for cell in case:
                # 判断是否为字符串类型，是则使用eval转换
                if isinstance(cell.value, str):
                    data.append(cell.value)
                else:
                    data.append(cell.value)
            data = [index] + data
            # [('',),('',),('',)] = case_data
            case = list(zip(titles, data))
            case_obj = Case(case)
            cases.append(case_obj)
        # 数据读完马上关闭表单，避免多个用例打开文件，最后回写数据不成功
        self.close()
        return cases

    def r_data_max(self, list1):
        """
        传入[]指定数据的列，读取每一行作为一个测试用例，放入字典中，所有用例放入列表返回
        :param list1:
        :return:
        """
        max_r = self.sheet.max_row  # 获取最大行
        # 定义一个空列表，存储所有测试用例
        cases = []
        titles = []  # 定义空列表存储表头
        for row in range(1, max_r + 1):
            if row == 1:  # 判断是否是第一行，如果是，则获取所有表头
                for column in list1:  # 遍历所有列
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
            else:  # 获取用例数据
                case_info = []  # 空列表存储该行的数据
                for column in list1:
                    info = self.sheet.cell(row, column).value
                    case_info.append(info)
                # 将表头和数据打包，放入字典中
                case = dict(list(zip(titles, case_info)))
                cases.append(case)
        return cases

    def r_data_max_obj(self, list1):
        # 读取数据之前打开表单
        self.open()
        max_r = self.sheet.max_row
        cases = []  # 存放用例
        titles = []  # 存放表头
        for row in range(1, max_r + 1):
            if row != 1:
                case_data = []
                for column in list1:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                case = list(zip(titles, case_data))
                case_obj = Case(case)
                for item in case:
                    setattr(case_obj, item[0], item[1])
                cases.append(case_obj)
            else:
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
        self.close()
        return cases

    # 最终读取excel的最优方法
    def readline_data_obj(self, list1=None):
        # 读取数据之前打开表单
        self.open()
        # 如果列表为空，则执行按行读取
        if len(list1) == 0:
            return self.read_data_obj()

        # 增加容错机制 例如输入大于最大列
        max_r = self.sheet.max_row
        cases = []  # 存放用例
        titles = []  # 存放表头
        for row in range(1, max_r + 1):
            if row != 1:
                case_data = []
                for column in list1:
                    info = self.sheet.cell(row, column).value
                    case_data.append(info)
                case = list(zip(titles, case_data))
                # 对下面代码的升级
                case_obj = Case(case)
                cases.append(case_obj)

                # case_obj = Case()
                # for item in case:
                #     setattr(case_obj, item[0], item[1])
                # cases.append(case_obj)
            else:
                for column in list1:
                    title = self.sheet.cell(row, column).value
                    titles.append(title)
                if None in titles:
                    raise ValueError('传入的列中，有表头为空的')
        self.close()
        return cases

    def write_list(self, row, column, msg):
        self.open()
        self.sheet.cell(row=1, column=1, value="销售商品ID")
        for item in eval(msg):
            # 写入数据 要保持save
            self.sheet.cell(row=row, column=column, value=item)
            row += 1
        self.wb.save(self.file_name)
        self.close()

    def write_data(self, row, column, msg):
        self.open()
        # 写入数据 要保持save
        self.sheet.cell(row=row, column=column, value=msg)
        self.wb.save(self.file_name)
        self.close()

    def write_sheet(self):
        self.open()
        # 写入数据 要保持save
        max_r = self.sheet.max_row
        for row in range(max_r):
            self.sheet.cell(row=row + 2, column=5, value=row + 2)
        self.wb.save(self.file_name)
        self.close()
        return "写入成功"

    def clear_sheet(self):
        self.open()
        # 写入数据 要保持save
        max_r = self.sheet.max_row
        for row in range(max_r):
            self.sheet.cell(row=row + 2, column=5, value="")
        # for row in range(len(li)):
        #     self.sheet.cell(row=max_r+row, column=1, value=li[row].用例名称)
        #     self.sheet.cell(row=max_r+row, column=3, value=li[row].接口地址)
        #     self.sheet.cell(row=max_r + row, column=6, value=li[row].请求数据)
        #     self.sheet.cell(row=max_r + row, column=7, value=li[row].期望返回状态)
        #     self.sheet.cell(row=max_r + row, column=9, value=li[row].期望返回数据)
        self.wb.save(self.file_name)
        self.close()
        return "删除成功"

    def extract_product_id(self):
        self.open()
        """
        按行读取数据 每个用例存储在一个对象中
        :return:
        """
        # 按行获取数据转化为列表
        rows_data = list(self.sheet.rows)
        product_id_list = []
        extend_id_list = []
        for data in rows_data:
            if data[1].value == "2":
                product_ids = list(eval(data[5].value)["product_data"].keys())
                product_id_list.extend(product_ids)
                if int(eval(data[5].value)["gift_data"]["status"]) == 1:
                    lis = eval(data[5].value)["gift_data"]["data"]["product_id"].split(",")
                    product_id_list.extend(lis)
                    extend_id_list.extend(lis)
        product_id_list = list(set(product_id_list))
        self.close()
        return extend_id_list, product_id_list

    def read_data_obj_xls(self, li):
        self.open()
        # 按行获取数据转化为列表
        rows_data = list(self.sheet.rows)
        # 获取表单的表头信息
        titles = [title.value for title in rows_data[0] if
                  title.value in ["用例名称", "接口地址", "是否登陆", "请求方式", "请求数据", "期望返回状态", "期望返回数据"]]
        # 定义一个空列表用来存储所有的用例
        cases = []
        flag = ""
        for index in range(len(rows_data[1:])):
            # 创建一个Case类的对象，用来保存用例数据
            case = rows_data[1:][index]
            # 参加促销活动 and 参与多倍积分测试案例 and 根据列表【2,4,6】编号筛选值
            condition = 'case[1].value == "1"' if len(
                li) == 0 else 'case[1].value == "1" and int(case[4].value) in li'  # and case[3].value == "1"
            if eval(condition):
                if case[0].value != None:
                    flag = str(case[0].value)
                data = [flag, case[2].value, case[3].value, case[4].value, case[5].value, case[6].value, case[8].value]
                case = list(zip(titles, data))
                case_obj = Case(case)
                cases.append(case_obj)
                # 数据读完马上关闭表单，避免多个用例打开文件，最后回写数据不成功
        self.close()
        return cases

    def get_excel_title(self, sheet_name):
        """获取sheet表头"""
        title_key = tuple(self.wb[sheet_name].iter_rows(max_row=1, values_only=True))[0]
        return title_key

    def get_all_value_2(self, sheet_name):
        """获取指定表单的所有数据(除去表头)"""
        rows_obj = self.wb[sheet_name].iter_rows(min_row=2, max_row=self.wb[sheet_name].max_row, values_only=True)
        values = []
        for row_tuple in rows_obj:
            value_list = []
            for value in row_tuple:
                value_list.append(value)
            values.append(value_list)
        return values

    def get_list_nametuple_all_value(self, sheet_name):
        """获取所有数据，返回嵌套命名元组的列表"""
        sheet_title = self.get_excel_title(sheet_name)
        values = self.get_all_value_2(sheet_name)
        excel = namedtuple('excel', sheet_title)
        value_list = []
        for value in values:
            e = excel(*value)
            value_list.append(e)
        return value_list


if __name__ == "__main__":

    wb = ReadExcel("FuTestDataSmoke_bk.xlsx",
                   'Sheet1')  # 无券促销测试   213 229
    res = wb.write_sheet()
    obj_list = wb.read_data_obj([])
    id_list = []

    for obj in obj_list:
        # except_return_data = eval(obj.期望返回数据)
        check_data = eval(obj.期望返回数据)["check_data"]
        info_order_confirm = eval(obj.期望返回数据)["cart_info_before_coupon"]
        if len(check_data.keys()) >= 2 and info_order_confirm["total_discount_amount"] > 0.00:
            # print(info_order_confirm["shipping_discount_amount"])
            # print(info_order_confirm["subtotal_discount_amount"])
            # print(info_order_confirm["discount_amount"])
            # print(info_order_confirm["tax_discount_amount"])
            # print(info_order_confirm["new_user_discount"])
            # print(info_order_confirm["use_coupon"]["code_save"])
            # print(info_order_confirm["total_discount_amount"])
            # id_list.append(obj.请求方式)
            print(obj.请求方式, len(check_data.keys()), check_data.keys())
            id_list.append(obj.请求方式)
    print(len(id_list))

    noCouponProduct = dict(noCouponProduct=id_list)
    print(noCouponProduct)
